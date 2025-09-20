import csv
import io
import os
import tempfile
import uuid
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models import Article, ArticleDocument

class CSVService:
    
    @staticmethod
    def validate_csv_file(file):
        if not file.filename.endswith('.csv'):
            raise ValueError('Formato de archivo inválido')
        
        content = file.stream.read()
        
        try:
            decoded_content = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                decoded_content = content.decode('utf-8')
            except UnicodeDecodeError:
                decoded_content = content.decode('UTF8')
        
        stream = io.StringIO(decoded_content, newline=None)
        csv_reader = csv.DictReader(stream)
        
        dois_in_csv = []
        for row in csv_reader:
            doi = row.get('DOI', '').strip()
            if doi:
                dois_in_csv.append(doi)
        
        existing_articles = []
        new_articles = []
        
        if dois_in_csv:
            existing_results = Article.check_multiple_dois(dois_in_csv)
            existing_dois = {row[0]: row[1] for row in existing_results}
            
            for doi in dois_in_csv:
                if doi in existing_dois:
                    existing_articles.append({
                        'doi': doi,
                        'titulo_original': existing_dois[doi]
                    })
                else:
                    new_articles.append({'doi': doi})
        
        # Formato de respuesta que espera el frontend
        if len(existing_articles) > 0:
            return {
                'status': 'duplicates_found',
                'total_in_csv': len(dois_in_csv),
                'existing_count': len(existing_articles),
                'new_count': len(new_articles),
                'existing_articles': existing_articles,
                'has_duplicates': True,
                'message': f'Se encontraron {len(existing_articles)} artículos que ya existen en la base de datos.'
            }
        else:
            return {
                'status': 'ready_to_import',
                'total_in_csv': len(dois_in_csv),
                'existing_count': len(existing_articles),
                'new_count': len(new_articles),
                'existing_articles': existing_articles,
                'has_duplicates': False,
                'message': f'Archivo validado correctamente. {len(new_articles)} artículos nuevos listos para importar.'
            }
    
    @staticmethod
    def import_csv_file(file, force_import=False):
        if not file.filename.endswith('.csv'):
            raise ValueError('Formato de archivo inválido')
        
        # Leer el contenido del archivo y manejar UTF-8 BOM
        content = file.stream.read()
        
        # Intentar decodificar como UTF-8 con BOM primero
        try:
            decoded_content = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            # Si falla, intentar UTF-8 normal
            try:
                decoded_content = content.decode('utf-8')
            except UnicodeDecodeError:
                # Como último recurso, intentar con la codificación original
                decoded_content = content.decode('UTF8')
        
        stream = io.StringIO(decoded_content, newline=None)
        csv_reader = csv.DictReader(stream)
        
        imported_count = 0
        skipped_count = 0
        
        for row in csv_reader:
            doi = row.get('DOI', '').strip()
            
            if not force_import and doi:
                if Article.check_doi_exists(doi):
                    skipped_count += 1
                    continue
            
            try:
                article_data = {
                    'autor': row.get('Authors', ''),
                    'nombre_revista': row.get('Source title', ''),
                    'anio': row.get('Year', ''),
                    'doi': doi,
                    'titulo_original': row.get('Title', ''),
                    'base_datos': 'Scopus',
                    'abstract': row.get('Abstract', ''),
                    'keywords_autor': row.get('Author Keywords', ''),
                    'keywords_indexed': row.get('Index Keywords', ''),
                    'enlace': row.get('Link', ''),
                    'eid': row.get('EID', '')
                }
                
                Article.create(article_data)
                imported_count += 1
                
            except Exception as e:
                print(f"Error importing row: {e}")
        
        message = f'{imported_count} artículos importados'
        if skipped_count > 0:
            message += f', {skipped_count} omitidos (ya existían)'
        
        return {
            'status': 'success',
            'message': message,
            'imported_count': imported_count,
            'skipped_count': skipped_count
        }

class ExcelService:
    # Configuración común para todas las exportaciones
    HEADERS = [
        'Autor', 'Nombre revista', 'Quartil revista', 'Fecha', 'DOI',
        'Título (original)', 'Título (español)', 'Base de datos', 'Abstract', 'Resumen',
        'Keywords autor', 'Keywords indexados', 'Problema a solucionar', 'Datos estadísticos',
        'Pregunta de investigación', 'Objetivo (original)', 'Objetivo (español)', 
        'Objetivo reescrito', 'Justificación', 'Hipótesis', 'Tipo investigación',
        'Estudios previos', 'Población/muestra/datos', 'Recolección datos',
        'Resultados', 'Conclusiones', 'Discusión', 'Trabajos futuros',
        'Enlace', 'EID', 'Seleccionado'
    ]
    
    COLUMN_WIDTHS = {
        1: 25,  # Autor
        2: 20,  # Revista
        3: 10,  # Quartil
        4: 8,   # Año
        5: 15,  # DOI
        6: 30,  # Título original
        7: 30,  # Título español
        8: 12,  # Base de datos
        9: 40,  # Abstract
        10: 40, # Resumen
        11: 20, # Keywords autor
        12: 20, # Keywords indexados
        13: 25, # Problema
        14: 20, # Datos estadísticos
        15: 25, # Pregunta investigación
        16: 25, # Objetivo original
        17: 25, # Objetivo español
        18: 25, # Objetivo reescrito
        19: 25, # Justificación
        20: 20, # Hipótesis
        21: 18, # Tipo investigación
        22: 25, # Estudios previos
        23: 25, # Población/muestra
        24: 25, # Recolección datos
        25: 30, # Resultados
        26: 30, # Conclusiones
        27: 30, # Discusión
        28: 25, # Trabajos futuros
        29: 25, # Enlace
        30: 15, # EID
        31: 12  # Seleccionado
    }
    
    # Columnas que requieren ajuste de texto (títulos, abstract, resumen)
    TEXT_WRAP_COLUMNS = [6, 7, 9, 10]
    
    @staticmethod
    def _get_local_timestamp():
        local_time = datetime.now()
        return local_time.strftime('%Y-%m-%d--%H-%M-%S')
    
    @staticmethod
    def _setup_header(ws, header_color='366092'):
        """Configura el encabezado del worksheet con estilo"""
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col, header in enumerate(ExcelService.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    @staticmethod
    def _populate_data(ws, articles):
        """Llena el worksheet con los datos de los artículos"""
        for row_idx, article in enumerate(articles, 2):
            for col_idx, value in enumerate(article, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Ajustar texto para campos largos
                if col_idx in ExcelService.TEXT_WRAP_COLUMNS:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    @staticmethod
    def _set_column_widths(ws):
        """Configura los anchos de columna"""
        for col, width in ExcelService.COLUMN_WIDTHS.items():
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    @staticmethod
    def _save_workbook(wb, base_filename):
        """Guarda el workbook y retorna el path y filename"""
        timestamp = ExcelService._get_local_timestamp()
        filename = f'{base_filename}-{timestamp}.xlsx'
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name, filename
    
    @staticmethod
    def _create_empty_bookmarks_sheet(ws):
        """Crea una hoja para marcadores vacía con mensaje informativo"""
        ws.title = "Marcadores (Vacío)"
        ExcelService._setup_header(ws)
        
        # Agregar mensaje informativo
        ws.cell(row=2, column=1, value="No hay artículos marcados como favoritos")
        ws.merge_cells('A2:AE2')
        
        # Centrar y estilizar el mensaje
        message_cell = ws.cell(row=2, column=1)
        message_cell.alignment = Alignment(horizontal='center', vertical='center')
        message_cell.font = Font(italic=True, color='666666')
    
    @staticmethod
    def create_excel_export():
        """Crea exportación Excel con todos los artículos"""
        articles = Article.get_all_for_export()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Análisis de Artículos"
        
        ExcelService._setup_header(ws)
        ExcelService._populate_data(ws, articles)
        ExcelService._set_column_widths(ws)
        
        return ExcelService._save_workbook(wb, 'matriz-analisis')
    
    @staticmethod
    def create_excel_export_bookmarks():
        """Crea exportación Excel solo con artículos marcados como favoritos"""
        articles = Article.get_bookmarks_for_export()
        
        wb = Workbook()
        ws = wb.active
        
        if not articles:
            # Crear hoja vacía con mensaje informativo
            ExcelService._create_empty_bookmarks_sheet(ws)
        else:
            # Crear hoja con datos de marcadores
            ws.title = "Marcadores - Análisis"
            ExcelService._setup_header(ws, header_color='1F4E79')  # Azul más oscuro para marcadores
            ExcelService._populate_data(ws, articles)
        
        ExcelService._set_column_widths(ws)
        return ExcelService._save_workbook(wb, 'matriz-marcadores')


class DocumentService:
    @staticmethod
    def get_upload_folder():
        current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(current_dir, 'uploads')
    
    ALLOWED_EXTENSIONS = {'pdf'}
    MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB
    
    @staticmethod
    def _create_upload_folder():
        upload_folder = DocumentService.get_upload_folder()
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder, exist_ok=True)
    
    @staticmethod
    def _allowed_file(filename):
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in DocumentService.ALLOWED_EXTENSIONS
    
    @staticmethod
    def _generate_unique_filename(original_filename):
        filename = secure_filename(original_filename)
        name, ext = os.path.splitext(filename)
        unique_name = f"{name}_{uuid.uuid4().hex}{ext}"
        return unique_name
    
    @staticmethod
    def upload_document(file, article_id, doc_type='original'):
        try:
            # Validaciones
            if not file or file.filename == '':
                raise ValueError('No se ha seleccionado ningún archivo')
            
            if not DocumentService._allowed_file(file.filename):
                raise ValueError('Solo se permiten archivos PDF')
            
            file.seek(0, os.SEEK_END)
            file_size = file.tell()
            file.seek(0)
            
            if file_size > DocumentService.MAX_FILE_SIZE:
                raise ValueError('El archivo es demasiado grande (máximo 16MB)')
            
            DocumentService._create_upload_folder()
            
            unique_filename = DocumentService._generate_unique_filename(file.filename)
            upload_folder = DocumentService.get_upload_folder()
            file_path = os.path.join(upload_folder, unique_filename)
            file.save(file_path)
            
            existing_docs = ArticleDocument.get_by_article_id(article_id)
            existing_doc = None
            
            for doc in existing_docs:
                doc_dict = ArticleDocument.to_dict(doc)
                if doc_type == 'original' and doc_dict['nombre_archivo_original']:
                    existing_doc = doc_dict
                    break
                elif doc_type == 'translated' and doc_dict['nombre_archivo_traducido']:
                    existing_doc = doc_dict
                    break
            
            if existing_doc:
                if doc_type == 'original':
                    old_file = existing_doc['nombre_archivo_original']
                    if old_file:
                        upload_folder = DocumentService.get_upload_folder()
                        old_path = os.path.join(upload_folder, old_file)
                        if os.path.exists(old_path):
                            os.remove(old_path)
                    
                    ArticleDocument.delete_by_article_and_type(article_id, 'original')
                    ArticleDocument.create(article_id, unique_filename, None)
                
                elif doc_type == 'translated':
                    old_file = existing_doc['nombre_archivo_traducido']
                    if old_file:
                        upload_folder = DocumentService.get_upload_folder()
                        old_path = os.path.join(upload_folder, old_file)
                        if os.path.exists(old_path):
                            os.remove(old_path)
                    
                    ArticleDocument.update_translated_filename(article_id, unique_filename)
            else:
                if doc_type == 'original':
                    ArticleDocument.create(article_id, unique_filename, None)
                elif doc_type == 'translated':
                    original_docs = [doc for doc in existing_docs 
                                   if ArticleDocument.to_dict(doc)['nombre_archivo_original']]
                    if original_docs:
                        ArticleDocument.update_translated_filename(article_id, unique_filename)
                    else:
                        ArticleDocument.create(article_id, None, unique_filename)
            
            return {
                'success': True,
                'message': f'Documento {doc_type} subido correctamente',
                'filename': unique_filename,
                'original_filename': file.filename
            }
            
        except Exception as e:
            try:
                if 'file_path' in locals() and os.path.exists(file_path):
                    os.remove(file_path)
            except:
                pass
            
            raise ValueError(str(e))
    
    @staticmethod
    def delete_document(article_id, doc_type):
        try:
            documents = ArticleDocument.get_by_article_id(article_id)
            
            filename_to_delete = None
            for doc in documents:
                doc_dict = ArticleDocument.to_dict(doc)
                if doc_type == 'original' and doc_dict['nombre_archivo_original']:
                    filename_to_delete = doc_dict['nombre_archivo_original']
                    break
                elif doc_type == 'translated' and doc_dict['nombre_archivo_traducido']:
                    filename_to_delete = doc_dict['nombre_archivo_traducido']
                    break
            
            if not filename_to_delete:
                raise ValueError('No se encontró el documento a eliminar')
            
            upload_folder = DocumentService.get_upload_folder()
            file_path = os.path.join(upload_folder, filename_to_delete)
            if os.path.exists(file_path):
                os.remove(file_path)
            
            # Eliminar de base de datos
            ArticleDocument.delete_by_article_and_type(article_id, doc_type)
            
            return {
                'success': True,
                'message': f'Documento {doc_type} eliminado correctamente'
            }
            
        except Exception as e:
            raise ValueError(str(e))
    
    @staticmethod
    def get_document_path(filename):
        upload_folder = DocumentService.get_upload_folder()
        return os.path.join(upload_folder, filename)